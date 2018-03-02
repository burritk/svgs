from openpyxl import Workbook
from openpyxl.utils import get_column_letter

with open('output.txt', 'r') as input_file:
    wb = Workbook()
    ws = wb.active
    row = 1
    ws[get_column_letter(0 + 1) + str(row)] = 'File'
    ws[get_column_letter(0 + 2) + str(row)] = 'Tags'
    row += 1
    for line in input_file:
        tags = ''
        for index, field in enumerate(line.split('"')):
            if index == 0:
                ws[get_column_letter(1) + str(row)] = field.strip()
            else:
                tags += field.strip() + ', '
        ws[get_column_letter(2) + str(row)] = tags[:-2]
        row += 1
    wb.save('tags.xlsx')